import pandas as pd
import streamlit as st
from pathlib import Path

EXCEL_PATH = Path.home() / "Library/CloudStorage/Dropbox/ScrivaniaMAC-Global/Websolute/__2026/Progetti Clienti/Clienti-Progetti attivi-Ennio-2026.xlsx"

RESP_ALIASES = {
    "Danile Garaffo": "Daniele Garaffo",
    "Caludio Tonti": "Claudio Tonti",
    "Enico Gualandi": "Enrico Gualandi",
}

SORT_COLS = {
    "—":              None,
    "Cliente":        "Cliente",
    "Resp. Progetto": "Resp.Progetto",
    "Attività":       "Attività",
    "Tipo Attività":  "Tipo Attività",
    "Stato Attività": "Stato Attività",
    "Rilascio":       "Data Rilascio",
    "TR/PR":          "TR_PR",
}

PM_EDITABLE_COLS = [
    "Tipo Attività",
    "Stato Attività",
    "Stato_Resp",
    "Note Ennio",
    "Next Step COM",
]

TIPO_OPTIONS = [
    "Presale Cliente",
    "Presale Prospect",
    "On-going",
    "Accettata",
]

STATO_OPTIONS = [
    "Critica",
    "Verifica Cliente",
    "Verifica PM",
    "Chiusa",
]

TIPO_COLORS = {
    "Presale Cliente":   "🟣",
    "Presale Prospect":  "🔵",
    "On-going":          "🟡",
    "Accettata":         "⚫",
}

STATO_COLORS = {
    "Critica":          "🔴",
    "Verifica Cliente": "🔵",
    "Verifica PM":      "🟠",
    "Chiusa":           "⚫",
}


@st.cache_data(ttl=5)
def load_data() -> pd.DataFrame:
    # Dopo il primo salvataggio il file non ha più la riga titolo — gestisci entrambi i casi
    df = pd.read_excel(EXCEL_PATH, header=0)
    if "Cliente" not in df.columns:
        df = pd.read_excel(EXCEL_PATH, header=1)
    df.rename(columns={
        "Stato - Attività Resp. Progetto": "Stato_Resp",
        "TR\nPR": "TR_PR",
        "Data Check  Rilascio": "Data Rilascio",
        "Data SAL interno": "Data SAL",
    }, inplace=True)
    df["Cliente"] = df["Cliente"].ffill()
    df["Resp.Progetto"] = df["Resp.Progetto"].replace(RESP_ALIASES)
    if "Tipo Attività" in df.columns:
        df["Tipo Attività"] = df["Tipo Attività"].replace({"Ongoing": "On-going"})
    if "Stato Attività" in df.columns:
        df["Stato Attività"] = df["Stato Attività"].replace({"Critico": "Critica"})
    for col in ["Stato_Resp", "Note Ennio", "Next Step COM", "Tipo Attività", "Stato Attività"]:
        if col in df.columns:
            df[col] = df[col].fillna("").str.strip()
    df["TR_PR"] = df["TR_PR"].fillna("").str.strip().replace("\xa0", "")
    df["Tipo_Icon"]  = df["Tipo Attività"].map(TIPO_COLORS).fillna("⚪")
    df["Stato_Icon"] = df["Stato Attività"].map(STATO_COLORS).fillna("⚪")
    df = df.dropna(subset=["Attività"])
    df = df.reset_index(drop=True)
    return df


def save_data(df: pd.DataFrame):
    import openpyxl
    import shutil
    from datetime import datetime

    st.cache_data.clear()

    # Backup automatico prima di sovrascrivere
    backup_dir = EXCEL_PATH.parent / "backup"
    backup_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    shutil.copy2(EXCEL_PATH, backup_dir / f"Clienti-Progetti_{ts}.xlsx")

    save_df = df.drop(columns=["Tipo_Icon", "Stato_Icon"], errors="ignore").copy()
    save_df.rename(columns={
        "Stato_Resp": "Stato - Attività Resp. Progetto",
        "TR_PR": "TR\nPR",
        "Data Rilascio": "Data Check  Rilascio",
        "Data SAL": "Data SAL interno",
    }, inplace=True)

    # Apre il file esistente — preserva tutta la formattazione Excel
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Trova la riga degli header
    header_row_idx = None
    col_map = {}  # nome colonna -> indice colonna 1-based
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if "Cliente" in row_vals:
            header_row_idx = r
            for c_idx, name in enumerate(row_vals, start=1):
                if name is not None:
                    col_map[str(name).strip()] = c_idx
            break

    if header_row_idx is None:
        save_df.to_excel(EXCEL_PATH, index=False, engine="openpyxl")
        return

    # Costruisce il lookup (cliente, attività) -> riga Excel, con forward-fill su Cliente
    cliente_col = col_map.get("Cliente", 1)
    attivita_col = col_map.get("Attività", 2)
    current_cliente = ""
    row_lookup = {}
    for r in range(header_row_idx + 1, ws.max_row + 1):
        raw_c = ws.cell(row=r, column=cliente_col).value
        raw_a = ws.cell(row=r, column=attivita_col).value
        if raw_c:
            current_cliente = str(raw_c).strip()
        attivita = str(raw_a or "").strip()
        if attivita:
            row_lookup[(current_cliente, attivita)] = r

    # Aggiorna solo i valori delle celle, lasciando intatta la formattazione
    for _, row in save_df.iterrows():
        cliente  = str(row.get("Cliente")  or "").strip()
        attivita = str(row.get("Attività") or "").strip()
        if not attivita:
            continue
        excel_r = row_lookup.get((cliente, attivita))
        if excel_r is not None:
            # Riga esistente: aggiorna celle
            for col_name, col_idx in col_map.items():
                if col_name in ("Cliente", "Attività"):
                    continue
                if col_name in save_df.columns:
                    val = row.get(col_name)
                    if pd.isna(val) or str(val) in ("nan", ""):
                        val = None
                    ws.cell(row=excel_r, column=col_idx).value = val
        else:
            # Riga nuova: aggiungi in fondo
            new_r = ws.max_row + 1
            for col_name, col_idx in col_map.items():
                if col_name in save_df.columns:
                    val = row.get(col_name)
                    if pd.isna(val) or str(val) in ("nan", ""):
                        val = None
                    ws.cell(row=new_r, column=col_idx).value = val

    wb.save(EXCEL_PATH)
