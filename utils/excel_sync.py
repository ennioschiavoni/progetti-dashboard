from pathlib import Path
import pandas as pd
import openpyxl

EXCEL_PATH = Path.home() / "Library/CloudStorage/Dropbox/ScrivaniaMAC-Global/Websolute/__2026/Progetti Clienti/Clienti-Progetti attivi-Ennio-2026.xlsx"

XLS_TO_CSV = {
    "Cliente":                            "Cliente",
    "Attività":                           "Attività",
    "Resp.Progetto":                      "Resp.Progetto",
    "Tipo Attività":                      "Tipo Attività",
    "Stato Attività":                     "Stato Attività",
    "Stato - Attività Resp. Progetto":    "Stato_Resp",
    "Note Ennio":                         "Note Ennio",
    "Next Step COM":                      "Next Step COM",
    "TR\nPR":                             "TR_PR",
    "Data Check  Rilascio":               "Data Rilascio",
    "Data SAL interno":                   "Data SAL",
    "Project Manager":                    "Project Manager",
    "Owner":                              "Owner",
}
CSV_TO_XLS = {v: k for k, v in XLS_TO_CSV.items()}


def excel_disponibile() -> bool:
    return EXCEL_PATH.exists()


def excel_to_df() -> pd.DataFrame:
    """Legge l'Excel e restituisce un DataFrame con i nomi colonne CSV."""
    df = pd.read_excel(EXCEL_PATH, sheet_name=0, dtype=str)
    df = df.rename(columns=XLS_TO_CSV)
    known = [c for c in XLS_TO_CSV.values() if c in df.columns]
    df = df[known].copy()
    df["Cliente"] = df["Cliente"].replace("", None).ffill()
    for col in df.columns:
        df[col] = df[col].fillna("").astype(str).str.strip()
        df[col] = df[col].replace({"nan": "", "None": "", "\xa0": ""})
    df = df[df["Attività"].str.strip() != ""].reset_index(drop=True)
    return df


def csv_to_excel(df: pd.DataFrame) -> tuple[int, int, list[str]]:
    """
    Aggiorna l'Excel con i dati del CSV preservando la formattazione.
    - Aggiorna le righe esistenti (match per Cliente+Attività)
    - Aggiunge in fondo le righe nuove (presenti nel CSV ma non in Excel)
    Ritorna (aggiornate, aggiunte, non_trovate_in_csv).
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Trova posizioni colonne dall'intestazione Excel (riga 1)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value is not None:
            headers[str(cell.value).strip()] = col_idx

    cliente_col = headers.get("Cliente")
    attivita_col = headers.get("Attività")
    if not cliente_col or not attivita_col:
        raise ValueError("Colonne Cliente/Attività non trovate nell'Excel")

    # Costruisce lookup (cliente, attività) → riga Excel
    lookup: dict[tuple, int] = {}
    current_cliente = ""
    for row_idx in range(2, ws.max_row + 2):
        c_val = ws.cell(row=row_idx, column=cliente_col).value
        a_val = ws.cell(row=row_idx, column=attivita_col).value
        if c_val:
            current_cliente = str(c_val).strip()
        if a_val:
            lookup[(current_cliente, str(a_val).strip())] = row_idx

    updated = 0
    added = 0
    next_row = ws.max_row + 1

    for _, row in df.iterrows():
        cliente  = str(row.get("Cliente", "")).strip()
        attivita = str(row.get("Attività", "")).strip()
        key = (cliente, attivita)

        if key in lookup:
            excel_row = lookup[key]
            for csv_col, xls_col in CSV_TO_XLS.items():
                if xls_col in headers and csv_col in df.columns:
                    val = row.get(csv_col, "")
                    ws.cell(row=excel_row, column=headers[xls_col]).value = val if val else None
            updated += 1
        else:
            # Riga nuova: aggiunge in fondo
            for csv_col, xls_col in CSV_TO_XLS.items():
                if xls_col in headers and csv_col in df.columns:
                    val = row.get(csv_col, "")
                    ws.cell(row=next_row, column=headers[xls_col]).value = val if val else None
            next_row += 1
            added += 1

    wb.save(EXCEL_PATH)
    return updated, added
